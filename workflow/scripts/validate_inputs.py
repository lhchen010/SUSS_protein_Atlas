"""Preflight input validation — run BEFORE the pipeline. Catches format problems
in seconds so a bad upload fails fast instead of 30 min into Foldseek. Portal calls
this on submit; CLI users can run it standalone:
    python workflow/scripts/validate_inputs.py config/config.yaml
Exit 0 = OK to run; exit 1 = blocking errors (printed). Warnings don't block.
"""
import os, sys, re, glob, json

def validate(cfg):
    errors, warns, info = [], [], {}
    strain = cfg.get("strain", {}).get("code", "")
    pdb_dir = cfg["input"]["pdb_dir"]
    accre = re.compile(r"[A-Z]{2,3}\d{4,}\.\d+")

    # ---- structures (required) ----
    pdbs = glob.glob(os.path.join(pdb_dir, "*.pdb"))
    info["n_structures"] = len(pdbs)
    if not pdbs:
        errors.append(f"no .pdb files in {pdb_dir}")
    bad_name, no_ca, lens = [], [], []
    accs = set()
    for p in pdbs[:100000]:
        b = os.path.basename(p)
        m = accre.search(b)
        if not m:
            bad_name.append(b)
        else:
            accs.add(m.group(0))
        ca = 0
        with open(p, encoding="utf-8", errors="replace") as fh:
            for line in fh:
                if line.startswith("ATOM") and line[12:16].strip() == "CA":
                    ca += 1
        if ca == 0: no_ca.append(b)
        else: lens.append(ca)
    if bad_name:
        errors.append(f"{len(bad_name)} PDB filenames lack a GenBank accession "
                      f"(need <strain>_<accession>.pdb): e.g. {bad_name[:3]}")
    if no_ca:
        errors.append(f"{len(no_ca)} PDBs have no CA atoms (empty/corrupt): e.g. {no_ca[:3]}")
    if lens:
        over = [l for l in lens if l > cfg['qc']['max_length']]
        if over: warns.append(f"{len(over)} structures exceed max_length "
                              f"{cfg['qc']['max_length']} aa (will fail QC)")
        info["length_range"] = [min(lens), max(lens)]
    if strain and pdbs and not any(os.path.basename(p).startswith(strain + "_") for p in pdbs[:50]):
        warns.append(f"no PDB filename starts with strain prefix '{strain}_' "
                     f"— provenance/merge may break")

    # ---- sequences (optional) ----
    seqs_fa = cfg["input"].get("seqs_fasta", "")
    if seqs_fa and os.path.exists(seqs_fa):
        seq_accs = set()
        n_seq = 0
        for line in open(seqs_fa, encoding="utf-8", errors="replace"):
            if line.startswith(">"):
                n_seq += 1
                m = accre.search(line)
                if m: seq_accs.add(m.group(0))
        info["n_sequences"] = n_seq
        missing = accs - seq_accs
        if missing:
            warns.append(f"{len(missing)} structures have no matching sequence "
                         f"(ESM/annotation will derive from structure): e.g. {sorted(missing)[:3]}")
    elif seqs_fa:
        warns.append(f"seqs_fasta '{seqs_fa}' not found — sequences derived from structures")

    # ---- RNAseq (optional, 2-sheet standard format) ----
    rna = cfg["input"].get("rnaseq_xlsx", "")
    if rna and os.path.exists(rna):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(rna, read_only=True)
            sheets = wb.sheetnames
            info["rnaseq_sheets"] = sheets
            need = {"id_mapping", "expression"}
            miss = need - set(sheets)
            if miss:
                errors.append(f"RNAseq xlsx missing required sheet(s): {sorted(miss)} "
                              f"(need id_mapping + expression) — see RNAseq_template")
            else:
                # id_mapping headers
                idm = wb["id_mapping"]
                hdr = [c.value for c in next(idm.iter_rows(max_row=1))]
                if not (hdr[:2] == ["protein_accession", "gene_id"]):
                    errors.append(f"id_mapping sheet headers must be "
                                  f"[protein_accession, gene_id], got {hdr[:2]}")
                else:
                    mapped = sum(1 for r in idm.iter_rows(min_row=2, values_only=True) if r and r[0])
                    info["rnaseq_mapped_accessions"] = mapped
                    exp = wb["expression"]
                    ehdr = [c.value for c in next(exp.iter_rows(max_row=1))]
                    if not ehdr or ehdr[0] != "gene_id":
                        errors.append("expression sheet first column must be 'gene_id'")
                    ncond = len([h for h in ehdr[1:] if h])
                    if ncond < 2:
                        warns.append(f"expression sheet has only {ncond} sample column(s); "
                                     f"need >=2 for condition grouping")
        except Exception as e:
            errors.append(f"cannot read RNAseq xlsx ({type(e).__name__}: {e})")
    elif rna:
        errors.append(f"rnaseq_xlsx '{rna}' set in config but file not found")
    else:
        info["rnaseq"] = "none — RNAseq step will be skipped"

    return errors, warns, info


if __name__ == "__main__":
    import yaml
    cfg_path = sys.argv[1] if len(sys.argv) > 1 else "config/config.yaml"
    cfg = yaml.safe_load(open(cfg_path))
    errors, warns, info = validate(cfg)
    print("=== SUSS input preflight ===")
    for k, v in info.items(): print(f"  {k}: {v}")
    for w in warns: print(f"  ⚠ WARN: {w}")
    for e in errors: print(f"  ✗ ERROR: {e}")
    if errors:
        print(f"\nBLOCKED — {len(errors)} error(s). Fix before running.")
        sys.exit(1)
    print(f"\nOK to run{' (' + str(len(warns)) + ' warning(s))' if warns else ''}.")
    sys.exit(0)
